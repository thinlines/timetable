from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
import sys
import re

font_size = 12
padding_inline = 5
line_spacing = 1.4

replacements = {
    "经济": "Econ",
    "统计": "Stat",
    "数学": "Math",
    "进阶": "Further ",
    "化学": "Chem",
    "艺术": "Art",
    "生物": "Bio",
    "物理": "Phys",
    "体育": "PE",
    "雅思": "IELTS",
    "自习不安排": "Study Hr. (No Tchr)",
    "自习": " Study Hr.",
    "语文": "Chinese",
    "外教": "FT ",
    "英语": "English",
    "美国历史": "US Hist",
    "英国历史": "British Hist",
    "IG": "IG ",
    "AP": "AP ",
    "AS": "AS ",
    "MED": " Med",
    "贾慧荣": " Zora",
    "胡捷": " Carol",
    "赵云飞": " Yunfei",
    "王岩": " Dawn",
    "侯静菲": " Jessica",
    "傅莹": " William",
    "郭艳": " Serena",
    "王京": " Ms. Wang Jing",
    "黄通昀": " Mr. Huang Tongyun",
    "封媛媛": " Ms. Feng Yuanyuan",
    "梁庆柱": " Mr. Liang Qingzhu",
    "刘蔚云": " Ms. Liu Weiyun",
    "陈改利": " Ms. Chen Gaili",
    "许英国": " Mr. Xu Yingguo",
    "屈晓瑜": "Ms. Qu Xiaoyu",
    "世界政治双语": "World Politics Bilingual",
    "世界地理双语": "World Geo Bilingual",
    "作息时间": "Period",
    "早读": "AM Study Hr.",
    "晚": "PM ",
    "升旗仪式": "Flag-Raising",
    "第一节": "1st",
    "第二节": "2nd",
    "第三节": "3rd",
    "第四节": "4th",
    "第五节": "5th",
    "第六节": "6th",
    "第七节": "7th",
    "第八节": "8th",
    "第九节": "9th",
    "高一": "G10.",
    "高二": "G11.",
    "高三": "G12.",
}


def reduce_spaces(text):
    text = text.strip()
    text = re.sub(" +", " ", text)
    text = re.sub("/", "\n", text)
    return text


def replace_text(cell_value, text, replacement):
    if isinstance(cell_value, str):  # Ensure it's a string
        new_value = cell_value.replace(text, replacement)
        new_value = reduce_spaces(new_value)
        return new_value
    return cell_value  # Return unchanged if not a string


def auto_adjust_dimensions(ws):
    """ Auto-adjust column widths and row heights based on cell content. """
    for col in ws.columns:  # Iterate over all columns
        max_length = max(
            max(len(line) for line in str(cell.value).split("\n"))
            if cell.value is not None else 0
            for cell in col if not isinstance(cell, MergedCell)
        )
        # Get column letter (A, B, C, ...)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + padding_inline

    for index, row in enumerate(ws.iter_rows(), start=1):
        max_height = max(len(str(cell.value).split("\n"))
                         if cell.value is not None else 1 for cell in row)
        print(index, max_height, max_height * font_size * line_spacing)
        ws.row_dimensions[index].hidden = False
        ws.row_dimensions[index].height = max_height * font_size * line_spacing


def main():
    # Load the workbook
    wb = load_workbook(sys.argv[1])

    # Select the worksheet
    ws = wb.active  # Or specify a sheet name: ws = wb["Sheet1"]

    for row in ws.iter_rows():
        for text, replacement in replacements.items():
            for cell in row:
                if cell.value:  # Only modify non-empty cells
                    cell.value = replace_text(cell.value, text, replacement)
                cell.font = Font(size=12, name="Calibri")

    auto_adjust_dimensions(ws)

    new_filename = ' '.join(["Modified", sys.argv[1]])
    wb.save(new_filename)
    wb.close()


if __name__ == "__main__":
    main()
